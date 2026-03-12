#!/usr/bin/env python3
"""
Process championship weekend data and append to existing season JSON files.

Game definitions live in championship_config.json — add a new season entry
there without touching this script.

Supported formats (set per season in championship_config.json):
  xlsx             — Multi-sheet xlsx (one sheet per team + Points sheet)
  stato_csv_folders — Stato CSV exports, one subfolder per team per game

Usage:
    python processing/process_championship.py
"""

import json
import pandas as pd
import openpyxl
from pathlib import Path
from typing import Any

from utils import safe_int, load_json, save_json, TEAM_NAMES

REPO_ROOT = Path(__file__).parent.parent
DATA_DIR = REPO_ROOT / "data"
CHAMP_DIR = REPO_ROOT / "input_data" / "championship"
CONFIG_PATH = Path(__file__).parent / "championship_config.json"

METERS_TO_YARDS = 1.09361

with open(CONFIG_PATH) as _f:
    CHAMP_CONFIG: dict[str, Any] = json.load(_f)


# ─── xlsx processing ──────────────────────────────────────────────────────────

def process_xlsx(filepath: Path, away: str, home: str, week: str) -> tuple[list[dict], list[dict]]:
    """Process a multi-sheet championship xlsx file."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    match_name = f"{away} @ {home}"

    player_rows: list[dict] = []
    team_summaries: list[dict] = []

    for team_abbrev in [away, home]:
        sheet_name = next(
            (s for s in wb.sheetnames if s.upper() == team_abbrev.upper()), None
        )
        if not sheet_name:
            print(f"  Warning: No sheet found for {team_abbrev} in {filepath.name}")
            continue

        ws = wb[sheet_name]
        headers = [cell.value for cell in ws[1]]

        team_goals = team_assists = team_blocks = team_turnovers = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            row_dict = dict(zip(headers, row))
            if not row_dict.get("Full Name"):
                continue

            jersey = str(row_dict.get("Jersey #", "")).replace(".0", "").strip()
            name = row_dict["Full Name"]
            player_name = f"{jersey.zfill(2)} {name}" if jersey else name

            goals     = safe_int(row_dict.get("Goals"))
            assists   = safe_int(row_dict.get("Assists"))
            blocks    = safe_int(row_dict.get("Blocks"))
            turnovers = safe_int(row_dict.get("Turnovers"))

            team_goals     += goals
            team_assists   += assists
            team_blocks    += blocks
            team_turnovers += turnovers

            player_rows.append({
                "player":     player_name,
                "team":       TEAM_NAMES.get(team_abbrev, team_abbrev),
                "teamAbbrev": team_abbrev,
                "week":       week,
                "match":      match_name,
                "goals":      goals,
                "assists":    assists,
                "blocks":     blocks,
                "turnovers":  turnovers,
                "+/-":        goals + assists + blocks - turnovers,
            })

        team_summaries.append({
            "team":     TEAM_NAMES.get(team_abbrev, team_abbrev),
            "abbrev":   team_abbrev,
            "match":    match_name,
            "week":     week,
            "goals":    team_goals,
            "blocks":   team_blocks,
            "turnovers": team_turnovers,
        })

    # Enrich from Points sheet if present
    points_ws = next(
        (wb[n] for n in ["Points", "points"] if n in wb.sheetnames), None
    )
    if points_ws:
        _enrich_team_stats_from_points(points_ws, team_summaries, away, home)

    return player_rows, team_summaries


def _enrich_team_stats_from_points(
    ws: Any, team_summaries: list[dict], away: str, home: str
) -> None:
    """Derive holds, break goals, and clean holds from the Points sheet."""
    points = []
    for row in ws.iter_rows(min_row=2, max_col=4, values_only=True):
        if row[1] and row[2]:
            points.append({
                "pulling": str(row[1]).strip(),
                "scoring": str(row[2]).strip(),
                "clean":   bool(row[3]) if row[3] is not None else False,
            })

    for summary in team_summaries:
        abbrev = summary["abbrev"]
        holds = clean_holds = break_goals = 0
        for pt in points:
            if pt["scoring"] == abbrev:
                if pt["pulling"] == abbrev:
                    break_goals += 1
                else:
                    holds += 1
                    if pt["clean"]:
                        clean_holds += 1
        summary["holds"] = holds
        summary["cleanHolds"] = clean_holds
        summary["breakGoals"] = break_goals


# ─── Stato CSV folder processing ──────────────────────────────────────────────

def process_stato_csv_folder(
    base_dir: Path, folder: str, away: str, home: str, week: str
) -> tuple[list[dict], list[dict]]:
    """Process a championship game folder containing Stato CSV exports."""
    game_dir = base_dir / folder
    if not game_dir.exists():
        print(f"  Warning: {game_dir} not found")
        return [], []

    match_name = f"{away} @ {home}"
    player_rows: list[dict] = []
    team_summaries: list[dict] = []

    col_renames = {
        "Player": "player",
        "Touches": "touches",
        "Throws": "throws",
        "Catches": "catches",
        "Defensive blocks": "blocks",
        "Goals": "goals",
        "Turnovers": "turnovers",
        "Assists": "assists",
        "Secondary assists": "secondaryAssists",
        "Offense points played": "offensePoints",
        "Defense points played": "defensePoints",
        "Possessions initiated": "possessionsInitiated",
        "Thrower errors": "throwerErrors",
        "Receiver errors": "receiverErrors",
        "Total completed throw gain (m)": "totalThrowYards",
        "Average completed throw gain (m)": "avgThrowYards",
        "Total caught pass gain (m)": "totalCatchYards",
        "Average caught pass gain (m)": "avgCatchYards",
    }

    for team_abbrev in [away, home]:
        team_dir = next(
            (d for d in game_dir.iterdir() if d.is_dir() and d.name.upper() == team_abbrev.upper()),
            None,
        )
        if not team_dir:
            print(f"  Warning: No folder for {team_abbrev} in {game_dir}")
            continue

        stats_csv = next(team_dir.glob("Player Stats*.csv"), None)
        if not stats_csv:
            print(f"  Warning: No Player Stats CSV for {team_abbrev}")
            continue

        df = pd.read_csv(stats_csv).rename(columns=col_renames)

        for col in ["totalThrowYards", "avgThrowYards", "totalCatchYards", "avgCatchYards"]:
            if col in df.columns:
                df[col] = (df[col] * METERS_TO_YARDS).round(1)

        df["team"]      = TEAM_NAMES.get(team_abbrev, team_abbrev)
        df["teamAbbrev"] = team_abbrev
        df["week"]      = week
        df["match"]     = match_name
        df["+/-"] = (
            df.get("goals", 0) + df.get("assists", 0) + df.get("blocks", 0)
            - df.get("turnovers", 0) - df.get("throwerErrors", 0) - df.get("receiverErrors", 0)
        )

        keep_cols = [
            "player", "team", "teamAbbrev", "week", "match",
            "goals", "assists", "secondaryAssists", "blocks", "turnovers",
            "touches", "throws", "catches",
            "offensePoints", "defensePoints", "possessionsInitiated",
            "throwerErrors", "receiverErrors",
            "totalThrowYards", "avgThrowYards", "totalCatchYards", "avgCatchYards",
            "+/-",
        ]
        df = df[[c for c in keep_cols if c in df.columns]].fillna(0)
        player_rows.extend(df.to_dict(orient="records"))

        team_summaries.append({
            "team":      TEAM_NAMES.get(team_abbrev, team_abbrev),
            "abbrev":    team_abbrev,
            "match":     match_name,
            "week":      week,
            "goals":     int(df["goals"].sum()),
            "blocks":    int(df["blocks"].sum()),
            "turnovers": int(df["turnovers"].sum()),
        })

    return player_rows, team_summaries


# ─── Main pipeline ────────────────────────────────────────────────────────────

def remove_championship_records(data: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Remove existing championship records so we can re-append cleanly."""
    return [r for r in data if r.get("week") not in ("Semifinals", "Finals")]


def process_season(year: int) -> None:
    """Process championship data for a given season using championship_config.json."""
    season_config = CHAMP_CONFIG.get(str(year))
    if not season_config:
        print(f"No championship config found for {year}")
        return

    champ_source = CHAMP_DIR / str(year)
    if not champ_source.exists():
        print(f"No championship data found for {year} at {champ_source}")
        return

    fmt = season_config["format"]
    games = season_config["games"]
    print(f"\nProcessing {year} championship data (format: {fmt})...")

    all_player_games: list[dict] = []
    all_team_games: list[dict] = []

    for game in games:
        away, home, week = game["away"], game["home"], game["week"]
        print(f"  Processing {week}: {away} @ {home}")

        if fmt == "xlsx":
            filepath = champ_source / game["file"]
            if not filepath.exists():
                print(f"  Warning: {filepath} not found, skipping")
                continue
            players, teams = process_xlsx(filepath, away, home, week)

        elif fmt == "stato_csv_folders":
            players, teams = process_stato_csv_folder(
                champ_source, game["folder"], away, home, week
            )

        else:
            print(f"  Unknown format '{fmt}', skipping")
            continue

        all_player_games.extend(players)
        all_team_games.extend(teams)

    if not all_player_games and not all_team_games:
        print(f"  No championship data processed for {year}")
        return

    data_dir = DATA_DIR / str(year)
    players_games_path = data_dir / "players_games.json"
    teams_games_path   = data_dir / "teams_games.json"

    existing_player_games = remove_championship_records(load_json(players_games_path))
    existing_team_games   = remove_championship_records(load_json(teams_games_path))

    existing_player_games.extend(all_player_games)
    existing_team_games.extend(all_team_games)

    save_json(existing_player_games, players_games_path)
    save_json(existing_team_games, teams_games_path)

    print(f"  {year} championship: {len(all_player_games)} player records, {len(all_team_games)} team records")


def main() -> None:
    for year in sorted(int(y) for y in CHAMP_CONFIG):
        process_season(year)
    print("\nDone!")


if __name__ == "__main__":
    main()
