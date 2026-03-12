#!/usr/bin/env python3
"""
Convert multi-sheet xlsx files to one CSV per sheet.

Run this as a pre-processing step when championship data arrives as xlsx files.
The resulting CSVs should be committed to the repo so the main pipeline
(process_championship.py) can run without an openpyxl dependency.

Usage:
    python processing/xlsx_to_csv.py input_data/championship/2025/
"""

import csv
import sys
import openpyxl
from pathlib import Path


def convert_xlsx_to_csvs(xlsx_path: Path, output_dir: Path) -> list[Path]:
    """Convert a multi-sheet xlsx to one CSV per sheet."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    output_dir.mkdir(parents=True, exist_ok=True)
    created = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        csv_path = output_dir / f"{xlsx_path.stem}_{sheet_name}.csv"
        with open(csv_path, "w", newline="") as f:
            writer = csv.writer(f)
            for row in ws.iter_rows(values_only=True):
                writer.writerow(row)
        print(f"  Wrote {csv_path.name}")
        created.append(csv_path)
    return created


def main() -> None:
    if len(sys.argv) < 2:
        print("Usage: python xlsx_to_csv.py <directory_or_xlsx_file>")
        sys.exit(1)

    target = Path(sys.argv[1])

    if target.is_file() and target.suffix == ".xlsx":
        xlsx_files = [target]
        output_dir = target.parent
    elif target.is_dir():
        xlsx_files = list(target.glob("*.xlsx"))
        output_dir = target
    else:
        print(f"Error: {target} is not an xlsx file or directory")
        sys.exit(1)

    if not xlsx_files:
        print(f"No xlsx files found in {target}")
        sys.exit(0)

    for xlsx_path in xlsx_files:
        print(f"\nConverting {xlsx_path.name}...")
        convert_xlsx_to_csvs(xlsx_path, output_dir)

    print("\nDone.")


if __name__ == "__main__":
    main()
